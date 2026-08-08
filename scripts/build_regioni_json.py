#!/usr/bin/env python3
"""
Converte "Mappatura competitors.xlsx" (Telemaco, foglio per regione) in
public/data/regioni.json, il formato consumato dalla pagina "Offerta" della
dashboard.

Uso:
    python3 scripts/build_regioni_json.py

Richiede openpyxl (`pip install openpyxl`). Rigenera public/data/regioni.json
ogni volta che lo Excel sorgente viene aggiornato con nuove regioni rilevate.
"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "Mappatura competitors.xlsx"
OUT_PATH = ROOT / "public" / "data" / "regioni.json"

FASCE_FATTURATO = [
    "< 10k €",
    "10-15k €",
    "15-25k €",
    "25-50k €",
    "50-75k €",
    "75-100k €",
    "100-150k €",
    "150-500k €",
    "0,5-1,5M €",
    "1,5-5M €",
    "> 5M €",
]

FONTE = {
    "studi": "Telemaco (codici ATECO 74.99.21 + 74.99.29, sedi attive)",
    "imprese": "InfoCamere/Movimprese, imprese attive CON DIPENDENTI, dato al 31/05/2026",
    "addetti": "InfoCamere/Movimprese, numero addetti (dipendenti + indipendenti) come media annua, dato al 31/05/2026",
}


def as_int(v):
    return None if v is None else int(v)


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Foglio2"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    # Foglio3: coorte 1990-99 (per ora sempre vuota, colonne predisposte per il futuro)
    coorte = {}
    if "Foglio3" in wb.sheetnames:
        for row in wb["Foglio3"].iter_rows(min_row=3, values_only=True):
            nome = row[0]
            if not nome or nome == "TOTALE ITALIA":
                continue
            coorte[nome] = {"studi_primario_9099": as_int(row[1]), "capitale_9099": as_int(row[2])}

    regioni_out = []
    for row in rows:
        nome = row[0]
        if not nome or nome == "TOTALE ITALIA":
            continue

        bands = [as_int(v) or 0 for v in row[1:12]]
        capitale = as_int(row[13])
        tot_completo = as_int(row[15])
        persona = as_int(row[16])
        prim_sec = as_int(row[18])
        sommerso = as_int(row[19])
        imprese = as_int(row[21])
        addetti = as_int(row[24])

        c9099 = coorte.get(nome, {"studi_primario_9099": None, "capitale_9099": None})

        regioni_out.append(
            {
                "nome": nome,
                "imprese_con_dipendenti": imprese,
                "numero_addetti": addetti,
                "studi_primario": tot_completo,
                "studi_prim_sec": prim_sec,
                "capitale": capitale,
                "persona": persona,
                "sommerso": sommerso,
                "bands": bands if tot_completo is not None else None,
                "studi_primario_9099": c9099["studi_primario_9099"],
                "capitale_9099": c9099["capitale_9099"],
            }
        )

    out = {"fonte": FONTE, "fasce_fatturato": FASCE_FATTURATO, "regioni": regioni_out}

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")

    rilevate = sum(1 for r in regioni_out if r["studi_primario"] is not None)
    print(f"Scritte {len(regioni_out)} regioni ({rilevate} rilevate) -> {OUT_PATH}")


if __name__ == "__main__":
    main()
